"""School Bus Fee Management — FastAPI backend (Phase 1 enhanced)."""
from __future__ import annotations

import io
import logging
import os
import random
import string
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import bcrypt
import jwt
from dotenv import load_dotenv
from fastapi import APIRouter, Depends, FastAPI, HTTPException, Query, status
from fastapi.responses import HTMLResponse, PlainTextResponse, Response
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr, Field
from starlette.middleware.cors import CORSMiddleware

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_EXPIRY_HOURS = int(os.environ.get("JWT_EXPIRY_HOURS", "720"))
SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "noreply@busfee.app")
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "kharwaramog02@gmail.com")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "12345678")

FIREBASE_CREDENTIALS_PATH = os.environ.get("FIREBASE_CREDENTIALS_PATH", "")
FIREBASE_BUCKET = os.environ.get("FIREBASE_BUCKET", "")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger("busfee")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

# ---------- Roles & Permissions ----------
ROLE_ADMIN = "admin"
ROLE_AUTHOR = "author"
ROLE_GUEST = "guest"
ALL_PAGES = ["dashboard", "schools", "students", "pending", "reports", "users"]
ROLE_DEFAULT_PERMS: Dict[str, List[str]] = {
    ROLE_ADMIN: ALL_PAGES,
    ROLE_AUTHOR: ["dashboard", "students", "pending", "reports"],
    ROLE_GUEST: ["dashboard"],  # admin can override per user
}
# Action permissions
ROLE_CAPS: Dict[str, Dict[str, bool]] = {
    ROLE_ADMIN: {"create": True, "edit": True, "delete": True, "export": True, "manage_users": True, "archive": True},
    ROLE_AUTHOR: {"create": True, "edit": True, "delete": False, "export": True, "manage_users": False, "archive": False},
    ROLE_GUEST: {"create": False, "edit": False, "delete": False, "export": False, "manage_users": False, "archive": False},
}
MAX_AUTHORS = 3

# Firebase Storage lazy init
_firebase_bucket = None
def _get_bucket():
    global _firebase_bucket
    if _firebase_bucket is not None:
        return _firebase_bucket
    if not FIREBASE_CREDENTIALS_PATH or not os.path.exists(FIREBASE_CREDENTIALS_PATH) or not FIREBASE_BUCKET:
        return None
    try:
        import firebase_admin
        from firebase_admin import credentials, storage as fb_storage
        if not firebase_admin._apps:
            cred = credentials.Certificate(FIREBASE_CREDENTIALS_PATH)
            firebase_admin.initialize_app(cred, {"storageBucket": FIREBASE_BUCKET})
        _firebase_bucket = fb_storage.bucket()
        return _firebase_bucket
    except Exception as e:
        logger.error(f"Firebase init failed: {e}")
        return None

app = FastAPI(title="School Bus Fee Management API")
api = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)


# ---------- Utility ----------
def hash_password(plain: str) -> str:
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def make_token(email: str, role: str = ROLE_ADMIN) -> str:
    return jwt.encode(
        {
            "sub": email,
            "role": role,
            "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRY_HOURS),
            "iat": datetime.now(timezone.utc),
        },
        JWT_SECRET,
        algorithm="HS256",
    )


def decode_token(token: str) -> Dict[str, Any]:
    return jwt.decode(token, JWT_SECRET, algorithms=["HS256"])


async def _admin_from_token(token: str) -> Dict[str, Any]:
    try:
        payload = decode_token(token)
    except jwt.PyJWTError:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"email": payload.get("sub")}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(401, "User not found")
    if user.get("status") and user["status"] != "active":
        raise HTTPException(403, "User inactive")
    return user


async def get_current_admin(
    creds: Optional[HTTPAuthorizationCredentials] = Depends(security),
    token: Optional[str] = Query(None),
) -> Dict[str, Any]:
    """Accept token from Authorization header OR ?token=... query param (for downloads)."""
    raw = creds.credentials if creds else token
    if not raw:
        raise HTTPException(401, "Missing token")
    return await _admin_from_token(raw)


def require_role(*roles: str):
    async def dep(user=Depends(get_current_admin)):
        if user.get("role") not in roles:
            raise HTTPException(403, "Insufficient permissions")
        return user
    return dep


def require_cap(cap: str):
    async def dep(user=Depends(get_current_admin)):
        caps = ROLE_CAPS.get(user.get("role", ""), {})
        if not caps.get(cap):
            raise HTTPException(403, f"Action '{cap}' not allowed for your role")
        return user
    return dep


def gen_otp() -> str:
    return "".join(random.choices(string.digits, k=6))


def send_email(to: str, subject: str, html: str) -> bool:
    if not SENDGRID_API_KEY:
        logger.warning("[DEV] No SENDGRID_API_KEY — printing email to console")
        logger.info(f"EMAIL TO: {to}\nSUBJECT: {subject}\nBODY:\n{html}")
        return True
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail
        message = Mail(from_email=SENDER_EMAIL, to_emails=to, subject=subject, html_content=html)
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        resp = sg.send(message)
        logger.info(f"SendGrid status {resp.status_code} to {to}")
        return resp.status_code in (200, 202)
    except Exception as e:
        logger.error(f"SendGrid error: {e}")
        return False


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------- Financial Year (Indian: April → March) ----------
def fy_label(dt: datetime) -> str:
    y = dt.year
    if dt.month < 4:
        return f"{y - 1}-{y}"
    return f"{y}-{y + 1}"


def fy_range(label: str) -> (datetime, datetime):
    """Return [start, end) of a financial year label like '2026-2027'."""
    try:
        a, b = label.split("-")
        start = datetime(int(a), 4, 1, tzinfo=timezone.utc)
        end = datetime(int(b), 4, 1, tzinfo=timezone.utc)
        return start, end
    except Exception:
        return datetime.min.replace(tzinfo=timezone.utc), datetime.max.replace(tzinfo=timezone.utc)


def _parse_dt(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        # accept both "YYYY-MM-DD" and ISO datetimes
        if "T" not in s and len(s) <= 10:
            d = datetime.fromisoformat(s)
        else:
            d = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if d.tzinfo is None:
            d = d.replace(tzinfo=timezone.utc)
        return d
    except Exception:
        return None


def _in_fy(value: Optional[str], fy: Optional[str]) -> bool:
    if not fy:
        return True
    dt = _parse_dt(value)
    if not dt:
        return False
    s, e = fy_range(fy)
    return s <= dt < e


# ---------- Models ----------
class LoginIn(BaseModel):
    email: EmailStr
    password: str


class TokenOut(BaseModel):
    token: str
    email: EmailStr


class ForgotIn(BaseModel):
    email: EmailStr


class VerifyResetIn(BaseModel):
    email: EmailStr
    otp: str
    new_password: str


class SchoolIn(BaseModel):
    name: str
    address: Optional[str] = ""
    contact_person: Optional[str] = ""
    contact_phone: Optional[str] = ""


class StudentIn(BaseModel):
    name: str
    parent_name: str
    parent_mobile: str
    school_id: str
    standard: str
    pickup_location: Optional[str] = ""
    yearly_fee: float
    admission_date: str  # ISO datetime
    due_date: str  # ISO datetime


class PaymentIn(BaseModel):
    amount: float
    payment_date: str  # ISO datetime
    mode: str  # cash / upi / bank
    note: Optional[str] = ""
    next_due_date: Optional[str] = None  # ISO datetime — next fee due


# ---------- Helpers ----------
def compute_status(yearly: float, paid: float) -> str:
    if paid <= 0:
        return "pending"
    if paid + 0.0001 >= yearly:
        return "completed"
    return "partial"


async def student_to_out(doc: Dict[str, Any]) -> Dict[str, Any]:
    school = await db.schools.find_one({"id": doc["school_id"]}, {"_id": 0, "name": 1})
    payments = await db.payments.find({"student_id": doc["id"]}, {"_id": 0}).to_list(1000)
    paid = sum(float(p["amount"]) for p in payments)
    yearly = float(doc["yearly_fee"])
    status_ = compute_status(yearly, paid)
    # last payment + next_due tracking
    last_payment_date: Optional[str] = None
    next_due: Optional[str] = doc.get("due_date")
    if payments:
        payments_sorted = sorted(
            payments, key=lambda p: _parse_dt(p.get("payment_date")) or datetime.min.replace(tzinfo=timezone.utc)
        )
        last = payments_sorted[-1]
        last_payment_date = last.get("payment_date")
        if last.get("next_due_date"):
            next_due = last["next_due_date"]
    overdue_days = 0
    nd = _parse_dt(next_due)
    if nd and status_ != "completed":
        delta = (datetime.now(timezone.utc) - nd).days
        overdue_days = max(delta, 0)
    return {
        **{k: v for k, v in doc.items() if k != "_id"},
        "school_name": (school or {}).get("name", "—"),
        "paid_amount": round(paid, 2),
        "pending_amount": round(max(yearly - paid, 0), 2),
        "status": status_,
        "last_payment_date": last_payment_date,
        "next_due_date": next_due,
        "overdue_days": overdue_days,
    }


# ---------- Startup ----------
@app.on_event("startup")
async def startup() -> None:
    await db.users.create_index("email", unique=True)
    await db.schools.create_index("id", unique=True)
    await db.students.create_index("id", unique=True)
    await db.payments.create_index("id", unique=True)
    await db.password_resets.create_index("email")
    await db.archives.create_index("fy", unique=True)

    # Migrate old admins → users
    async for old in db.admins.find({}):
        await db.users.update_one(
            {"email": old["email"]},
            {"$setOnInsert": {
                "id": str(uuid.uuid4()),
                "email": old["email"],
                "password_hash": old["password_hash"],
                "full_name": old.get("email", "").split("@")[0].title(),
                "mobile": "",
                "role": ROLE_ADMIN,
                "status": "active",
                "page_permissions": ALL_PAGES,
                "created_at": old.get("created_at", now_iso()),
                "last_login": None,
            }},
            upsert=True,
        )
    await db.admins.drop()

    # Enforce single admin = ADMIN_EMAIL.  Remove any users that aren't ADMIN_EMAIL and have role admin without legit reason.
    # Seed admin if missing
    existing = await db.users.find_one({"email": ADMIN_EMAIL})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": ADMIN_EMAIL,
            "password_hash": hash_password(ADMIN_PASSWORD),
            "full_name": "Bus Fee Admin",
            "mobile": "",
            "role": ROLE_ADMIN,
            "status": "active",
            "page_permissions": ALL_PAGES,
            "created_at": now_iso(),
            "last_login": None,
        })
        logger.info(f"Seeded admin: {ADMIN_EMAIL}")
    else:
        # ensure admin is active + has full perms
        await db.users.update_one(
            {"email": ADMIN_EMAIL},
            {"$set": {"role": ROLE_ADMIN, "status": "active", "page_permissions": ALL_PAGES}},
        )

    if await db.schools.count_documents({}) == 0:
        sample_schools = [
            {"id": str(uuid.uuid4()), "name": "JB School", "address": "Main Road, Anand", "contact_person": "Mr. Joshi", "contact_phone": "9876500011", "created_at": now_iso()},
            {"id": str(uuid.uuid4()), "name": "DP School", "address": "Station Road, Vadodara", "contact_person": "Mrs. Patel", "contact_phone": "9876500022", "created_at": now_iso()},
        ]
        await db.schools.insert_many([dict(s) for s in sample_schools])
        logger.info("Seeded sample schools")


@app.on_event("shutdown")
async def shutdown() -> None:
    client.close()


# ---------- Auth ----------
@api.post("/auth/login", response_model=TokenOut)
async def login(body: LoginIn):
    user = await db.users.find_one({"email": body.email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(401, "Invalid email or password")
    if user.get("status") and user["status"] != "active":
        raise HTTPException(403, "User is inactive")
    await db.users.update_one({"email": body.email}, {"$set": {"last_login": now_iso()}})
    return TokenOut(token=make_token(body.email, user.get("role", ROLE_ADMIN)), email=body.email)


@api.get("/auth/me")
async def me(user=Depends(get_current_admin)):
    role = user.get("role", ROLE_ADMIN)
    return {
        **{k: v for k, v in user.items() if k != "password_hash"},
        "capabilities": ROLE_CAPS.get(role, {}),
    }


@api.post("/auth/forgot-password")
async def forgot_password(body: ForgotIn):
    user = await db.users.find_one({"email": body.email})
    if user:
        otp = gen_otp()
        await db.password_resets.update_one(
            {"email": body.email},
            {"$set": {
                "email": body.email, "otp": otp,
                "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat(),
                "consumed": False, "created_at": now_iso(),
            }},
            upsert=True,
        )
        html = f"""<div style='font-family:system-ui,sans-serif;padding:24px;background:#f9fafb;'>
          <h2 style='color:#2B4C3E;'>Bus Fee Manager — Password Reset Approval</h2>
          <p>We received a request to reset the password for <b>{body.email}</b>.</p>
          <p>Your one-time approval code (valid for 15 minutes) is:</p>
          <div style='font-size:34px;font-weight:700;letter-spacing:6px;background:#fff;padding:18px 24px;border-radius:12px;display:inline-block;border:1px solid #E5E7EB;color:#111827;'>{otp}</div>
          <p style='margin-top:24px;'>Enter this code in the app to approve and set your new password.</p>
        </div>"""
        send_email(body.email, "Approve Password Reset — Bus Fee Manager", html)
    return {"ok": True, "message": "If the email exists, an OTP has been sent."}


@api.post("/auth/verify-reset")
async def verify_reset(body: VerifyResetIn):
    rec = await db.password_resets.find_one({"email": body.email})
    if not rec or rec.get("consumed"):
        raise HTTPException(400, "Invalid or already used code")
    if rec["otp"] != body.otp:
        raise HTTPException(400, "Incorrect OTP")
    try:
        exp = datetime.fromisoformat(rec["expires_at"])
    except Exception:
        raise HTTPException(400, "Invalid OTP record")
    if exp < datetime.now(timezone.utc):
        raise HTTPException(400, "OTP expired")
    if len(body.new_password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")
    await db.users.update_one({"email": body.email}, {"$set": {"password_hash": hash_password(body.new_password)}})
    await db.password_resets.update_one({"email": body.email}, {"$set": {"consumed": True}})
    return {"ok": True, "message": "Password approved and updated."}


# ---------- Schools ----------
@api.get("/schools")
async def list_schools(admin=Depends(get_current_admin)):
    schools = await db.schools.find({}, {"_id": 0}).to_list(1000)
    out = []
    for s in schools:
        cnt = await db.students.count_documents({"school_id": s["id"]})
        s["student_count"] = cnt
        out.append(s)
    return out


@api.post("/schools")
async def create_school(body: SchoolIn, admin=Depends(get_current_admin)):
    doc = {**body.model_dump(), "id": str(uuid.uuid4()), "created_at": now_iso()}
    await db.schools.insert_one(dict(doc))
    return {k: v for k, v in doc.items() if k != "_id"}


@api.get("/schools/{school_id}")
async def get_school(school_id: str, admin=Depends(get_current_admin)):
    s = await db.schools.find_one({"id": school_id}, {"_id": 0})
    if not s:
        raise HTTPException(404, "School not found")
    return s


@api.put("/schools/{school_id}")
async def update_school(school_id: str, body: SchoolIn, admin=Depends(get_current_admin)):
    res = await db.schools.update_one({"id": school_id}, {"$set": body.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "School not found")
    return await db.schools.find_one({"id": school_id}, {"_id": 0})


@api.delete("/schools/{school_id}")
async def delete_school(school_id: str, admin=Depends(get_current_admin)):
    students = await db.students.find({"school_id": school_id}, {"id": 1, "_id": 0}).to_list(10000)
    sids = [s["id"] for s in students]
    if sids:
        await db.payments.delete_many({"student_id": {"$in": sids}})
        await db.students.delete_many({"school_id": school_id})
    await db.schools.delete_one({"id": school_id})
    return {"ok": True}


# ---------- Students ----------
@api.get("/students")
async def list_students(
    school_id: Optional[str] = None,
    search: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    standard: Optional[str] = None,
    due: Optional[str] = None,
    fy: Optional[str] = None,
    admin=Depends(get_current_admin),
):
    query: Dict[str, Any] = {}
    if school_id:
        query["school_id"] = school_id
    if standard:
        query["standard"] = standard
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"parent_name": {"$regex": search, "$options": "i"}},
            {"parent_mobile": {"$regex": search, "$options": "i"}},
        ]
    docs = await db.students.find(query, {"_id": 0}).to_list(5000)
    if fy:
        docs = [d for d in docs if _in_fy(d.get("admission_date"), fy)]
    out = [await student_to_out(d) for d in docs]
    if status_filter and status_filter in ("pending", "partial", "completed"):
        out = [s for s in out if s["status"] == status_filter]
    if due:
        today = datetime.now(timezone.utc).date()
        end = today + (timedelta(days=0) if due == "today" else timedelta(days=7))
        kept = []
        for s in out:
            d = _parse_dt(s.get("next_due_date") or s.get("due_date"))
            if not d:
                continue
            if due == "today" and d.date() == today:
                kept.append(s)
            elif due == "week" and today <= d.date() <= end:
                kept.append(s)
        out = kept
    return out


@api.post("/students")
async def create_student(body: StudentIn, admin=Depends(get_current_admin)):
    school = await db.schools.find_one({"id": body.school_id}, {"_id": 0})
    if not school:
        raise HTTPException(400, "Invalid school_id")
    doc = {**body.model_dump(), "id": str(uuid.uuid4()), "created_at": now_iso()}
    await db.students.insert_one(dict(doc))
    return await student_to_out(doc)


@api.get("/students/{student_id}")
async def get_student(student_id: str, admin=Depends(get_current_admin)):
    s = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not s:
        raise HTTPException(404, "Student not found")
    return await student_to_out(s)


@api.put("/students/{student_id}")
async def update_student(student_id: str, body: StudentIn, admin=Depends(get_current_admin)):
    school = await db.schools.find_one({"id": body.school_id}, {"_id": 0})
    if not school:
        raise HTTPException(400, "Invalid school_id")
    res = await db.students.update_one({"id": student_id}, {"$set": body.model_dump()})
    if res.matched_count == 0:
        raise HTTPException(404, "Student not found")
    doc = await db.students.find_one({"id": student_id}, {"_id": 0})
    return await student_to_out(doc)


@api.delete("/students/{student_id}")
async def delete_student(student_id: str, admin=Depends(get_current_admin)):
    await db.payments.delete_many({"student_id": student_id})
    await db.students.delete_one({"id": student_id})
    return {"ok": True}


# ---------- Payments ----------
@api.get("/students/{student_id}/payments")
async def list_payments(student_id: str, admin=Depends(get_current_admin)):
    docs = await db.payments.find({"student_id": student_id}, {"_id": 0}).sort("payment_date", -1).to_list(1000)
    return docs


@api.post("/students/{student_id}/payments")
async def add_payment(student_id: str, body: PaymentIn, admin=Depends(get_current_admin)):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(404, "Student not found")
    if body.amount <= 0:
        raise HTTPException(400, "Amount must be positive")
    doc = {
        **body.model_dump(),
        "id": str(uuid.uuid4()),
        "student_id": student_id,
        "created_at": now_iso(),
    }
    await db.payments.insert_one(dict(doc))
    return {k: v for k, v in doc.items() if k != "_id"}


@api.delete("/payments/{payment_id}")
async def delete_payment(payment_id: str, admin=Depends(get_current_admin)):
    await db.payments.delete_one({"id": payment_id})
    return {"ok": True}


# ---------- Pending Fees ----------
@api.get("/pending-fees")
async def pending_fees(fy: Optional[str] = None, admin=Depends(get_current_admin)):
    """Return students whose next_due_date is past today and balance > 0."""
    docs = await db.students.find({}, {"_id": 0}).to_list(10000)
    if fy:
        docs = [d for d in docs if _in_fy(d.get("admission_date"), fy)]
    out = []
    today = datetime.now(timezone.utc)
    for d in docs:
        full = await student_to_out(d)
        if full["status"] == "completed":
            continue
        nd = _parse_dt(full.get("next_due_date") or full.get("due_date"))
        if not nd or nd > today:
            continue
        out.append(full)
    out.sort(key=lambda s: -s["overdue_days"])
    return out


# ---------- Dashboard ----------
@api.get("/dashboard/summary")
async def dashboard_summary(fy: Optional[str] = None, admin=Depends(get_current_admin)):
    docs = await db.students.find({}, {"_id": 0}).to_list(10000)
    if fy:
        docs = [d for d in docs if _in_fy(d.get("admission_date"), fy)]
    total_schools = await db.schools.count_documents({})
    total_students = len(docs)
    total_yearly = sum(float(d["yearly_fee"]) for d in docs)
    sids = [d["id"] for d in docs]
    pay_q: Dict[str, Any] = {}
    if sids:
        pay_q["student_id"] = {"$in": sids}
    payments = await db.payments.find(pay_q, {"_id": 0}).to_list(50000)
    if fy:
        payments = [p for p in payments if _in_fy(p.get("payment_date"), fy)]
    total_collected = sum(float(p["amount"]) for p in payments)
    # completed: students fully paid
    by_student: Dict[str, float] = {}
    for p in payments:
        by_student[p["student_id"]] = by_student.get(p["student_id"], 0.0) + float(p["amount"])
    completed_total = sum(float(d["yearly_fee"]) for d in docs if by_student.get(d["id"], 0.0) + 0.0001 >= float(d["yearly_fee"]))
    pending_total = sum(max(float(d["yearly_fee"]) - by_student.get(d["id"], 0.0), 0) for d in docs)
    return {
        "total_schools": total_schools,
        "total_students": total_students,
        "total_yearly": round(total_yearly, 2),
        "total_collected": round(total_collected, 2),
        "total_pending": round(pending_total, 2),
        "total_completed": round(completed_total, 2),
    }


@api.get("/dashboard/by-school")
async def dashboard_by_school(fy: Optional[str] = None, admin=Depends(get_current_admin)):
    schools = await db.schools.find({}, {"_id": 0}).to_list(1000)
    out = []
    for s in schools:
        students = await db.students.find({"school_id": s["id"]}, {"_id": 0}).to_list(5000)
        if fy:
            students = [st for st in students if _in_fy(st.get("admission_date"), fy)]
        sids = [st["id"] for st in students]
        yearly = sum(float(st["yearly_fee"]) for st in students)
        collected = 0.0
        if sids:
            pq = await db.payments.find({"student_id": {"$in": sids}}, {"_id": 0}).to_list(50000)
            if fy:
                pq = [p for p in pq if _in_fy(p.get("payment_date"), fy)]
            collected = sum(float(p["amount"]) for p in pq)
        out.append({
            "school_id": s["id"], "school_name": s["name"], "student_count": len(students),
            "yearly_total": round(yearly, 2), "collected": round(collected, 2),
            "pending": round(max(yearly - collected, 0), 2),
        })
    return out


# ---------- Financial Years ----------
@api.get("/financial-years")
async def financial_years(admin=Depends(get_current_admin)):
    """List FY labels derived from admission/payment dates, plus current and next."""
    years: set = set()
    cur = datetime.now(timezone.utc)
    years.add(fy_label(cur))
    # next year
    next_fy = datetime(cur.year + 1 if cur.month >= 4 else cur.year, 4, 1, tzinfo=timezone.utc)
    years.add(fy_label(next_fy))
    # previous year
    prev = datetime(cur.year - 1 if cur.month < 4 else cur.year, 3, 1, tzinfo=timezone.utc)
    years.add(fy_label(prev))
    for coll, field in [("students", "admission_date"), ("payments", "payment_date")]:
        async for doc in db[coll].find({}, {field: 1, "_id": 0}):
            dt = _parse_dt(doc.get(field))
            if dt:
                years.add(fy_label(dt))
    sorted_years = sorted(years, key=lambda x: int(x.split("-")[0]), reverse=True)
    return {"current": fy_label(cur), "years": sorted_years}


# ---------- Reports: real PDF + real Excel ----------
def _format_inr(n: float) -> str:
    s = f"{int(round(n)):,}"
    parts = s.split(",")
    if len(parts) > 1:
        first, rest = parts[0], "".join(parts[1:])
        groups = []
        while len(rest) > 3:
            groups.insert(0, rest[-3:]); rest = rest[:-3]
        groups.insert(0, rest)
        head: List[str] = []
        while len(first) > 2:
            head.insert(0, first[-2:]); first = first[:-2]
        if first:
            head.insert(0, first)
        return "Rs. " + ",".join(head + groups)
    return "Rs. " + s


def _fmt_dt(s: Optional[str]) -> str:
    if not s:
        return "—"
    dt = _parse_dt(s)
    if not dt:
        return s
    return dt.strftime("%d/%m/%Y %H:%M")


async def _gather_report_rows(
    school_id: Optional[str], status_filter: Optional[str], start: Optional[str], end: Optional[str], fy: Optional[str]
):
    query: Dict[str, Any] = {}
    if school_id:
        query["school_id"] = school_id
    docs = await db.students.find(query, {"_id": 0}).to_list(10000)
    if fy:
        docs = [d for d in docs if _in_fy(d.get("admission_date"), fy)]
    out = []
    s_dt = _parse_dt(start)
    e_dt = _parse_dt(end)
    for d in docs:
        s = await student_to_out(d)
        if status_filter and s["status"] != status_filter:
            continue
        nd = _parse_dt(s.get("next_due_date") or s.get("due_date"))
        if s_dt and nd and nd < s_dt:
            continue
        if e_dt and nd and nd > e_dt:
            continue
        out.append(s)
    return out


@api.get("/reports/excel")
async def report_excel(
    school_id: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    fy: Optional[str] = None,
    admin=Depends(get_current_admin),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = await _gather_report_rows(school_id, status_filter, start, end, fy)
    wb = Workbook()
    ws = wb.active
    ws.title = "Bus Fee Report"

    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2B4C3E")
    head_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    money = Font(name="Calibri", size=10)

    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = f"Bus Fee Report — Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cell.font = title_font
    cell.fill = head_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Student", "School", "Class", "Parent", "Mobile", "Yearly Fee", "Paid", "Pending", "Status", "Next Due"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")

    total_y = total_p = total_pend = 0.0
    for idx, s in enumerate(rows, start=4):
        ws.cell(row=idx, column=1, value=s["name"])
        ws.cell(row=idx, column=2, value=s["school_name"])
        ws.cell(row=idx, column=3, value=s["standard"])
        ws.cell(row=idx, column=4, value=s["parent_name"])
        ws.cell(row=idx, column=5, value=s["parent_mobile"])
        ws.cell(row=idx, column=6, value=float(s["yearly_fee"])).number_format = '"₹"#,##0'
        ws.cell(row=idx, column=7, value=float(s["paid_amount"])).number_format = '"₹"#,##0'
        ws.cell(row=idx, column=8, value=float(s["pending_amount"])).number_format = '"₹"#,##0'
        ws.cell(row=idx, column=9, value=s["status"].title())
        ws.cell(row=idx, column=10, value=_fmt_dt(s.get("next_due_date")))
        total_y += s["yearly_fee"]
        total_p += s["paid_amount"]
        total_pend += s["pending_amount"]

    tot_row = len(rows) + 5
    ws.cell(row=tot_row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tot_row, column=6, value=total_y).number_format = '"₹"#,##0'
    ws.cell(row=tot_row, column=7, value=total_p).number_format = '"₹"#,##0'
    ws.cell(row=tot_row, column=8, value=total_pend).number_format = '"₹"#,##0'
    for col in range(5, 9):
        ws.cell(row=tot_row, column=col).font = Font(bold=True)

    widths = [22, 22, 8, 22, 14, 14, 14, 14, 12, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"bus-fee-report-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.get("/reports/pdf")
async def report_pdf(
    school_id: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    fy: Optional[str] = None,
    admin=Depends(get_current_admin),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    rows = await _gather_report_rows(school_id, status_filter, start, end, fy)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#2B4C3E"))
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#6B7280"))

    story: List[Any] = []
    story.append(Paragraph("Bus Fee Management Report", title_style))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        f"{' · FY: ' + fy if fy else ''}"
        f"{' · Status: ' + status_filter.title() if status_filter else ''}",
        sub_style,
    ))
    story.append(Spacer(1, 10))

    headers = ["Student", "School", "Class", "Parent", "Mobile", "Yearly", "Paid", "Pending", "Status", "Next Due"]
    data: List[List[Any]] = [headers]
    total_y = total_p = total_pend = 0.0
    for s in rows:
        data.append([
            s["name"], s["school_name"], s["standard"], s["parent_name"], s["parent_mobile"],
            _format_inr(s["yearly_fee"]), _format_inr(s["paid_amount"]), _format_inr(s["pending_amount"]),
            s["status"].title(), _fmt_dt(s.get("next_due_date")),
        ])
        total_y += s["yearly_fee"]
        total_p += s["paid_amount"]
        total_pend += s["pending_amount"]
    data.append(["", "", "", "", "TOTAL", _format_inr(total_y), _format_inr(total_p), _format_inr(total_pend), "", ""])

    tbl = Table(data, repeatRows=1, colWidths=[28*mm, 28*mm, 14*mm, 28*mm, 24*mm, 22*mm, 22*mm, 22*mm, 18*mm, 26*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B4C3E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (5, 0), (7, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -2), 0.4, colors.HexColor("#E5E7EB")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F2F5F3")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor("#2B4C3E")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    fname = f"bus-fee-report-{datetime.now().strftime('%Y%m%d-%H%M')}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'},
    )


# Keep legacy endpoints
@api.get("/reports/csv", response_class=PlainTextResponse)
async def report_csv(
    school_id: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    fy: Optional[str] = None,
    admin=Depends(get_current_admin),
):
    rows = await _gather_report_rows(school_id, status_filter, start, end, fy)
    out = ["Student,School,Class,Parent,Mobile,Yearly Fee,Paid,Pending,Status,Next Due"]
    for s in rows:
        out.append(
            f'"{s["name"]}","{s["school_name"]}","{s["standard"]}","{s["parent_name"]}",'
            f'"{s["parent_mobile"]}",{s["yearly_fee"]},{s["paid_amount"]},{s["pending_amount"]},'
            f'{s["status"]},"{_fmt_dt(s.get("next_due_date"))}"'
        )
    return "\n".join(out)


@api.get("/reports/html", response_class=HTMLResponse)
async def report_html(
    school_id: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    fy: Optional[str] = None,
    admin=Depends(get_current_admin),
):
    rows = await _gather_report_rows(school_id, status_filter, start, end, fy)
    body_rows: List[str] = []
    total_y = total_p = total_pend = 0.0
    for s in rows:
        body_rows.append(
            f"<tr><td>{s['name']}</td><td>{s['school_name']}</td><td>{s['standard']}</td>"
            f"<td>{_format_inr(s['yearly_fee'])}</td><td>{_format_inr(s['paid_amount'])}</td>"
            f"<td>{_format_inr(s['pending_amount'])}</td><td>{s['status']}</td>"
            f"<td>{_fmt_dt(s.get('next_due_date'))}</td></tr>"
        )
        total_y += s["yearly_fee"]; total_p += s["paid_amount"]; total_pend += s["pending_amount"]
    return f"""<html><head><meta charset='utf-8'><title>Bus Fee Report</title>
    <style>body{{font-family:system-ui;padding:24px;color:#111}}h1{{color:#2B4C3E}}
    table{{width:100%;border-collapse:collapse}}th,td{{padding:8px;border-bottom:1px solid #E5E7EB;text-align:left;font-size:13px}}
    th{{background:#F3F4F6}}.tot{{margin-top:16px;background:#F2F5F3;padding:12px;border-radius:8px}}</style></head>
    <body><h1>Bus Fee Report</h1><p>Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
    <table><thead><tr><th>Student</th><th>School</th><th>Class</th><th>Yearly</th><th>Paid</th><th>Pending</th><th>Status</th><th>Next Due</th></tr></thead>
    <tbody>{''.join(body_rows) or '<tr><td colspan=8>No records</td></tr>'}</tbody></table>
    <div class='tot'><b>Total Yearly:</b> {_format_inr(total_y)} &nbsp; <b>Total Collected:</b> {_format_inr(total_p)} &nbsp; <b>Total Pending:</b> {_format_inr(total_pend)}</div>
    </body></html>"""


# ---------- User Management (Admin only) ----------
class UserIn(BaseModel):
    full_name: str
    email: EmailStr
    mobile: Optional[str] = ""
    password: str
    role: str = ROLE_GUEST
    status: str = "active"
    page_permissions: Optional[List[str]] = None


class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    mobile: Optional[str] = None
    role: Optional[str] = None
    status: Optional[str] = None
    page_permissions: Optional[List[str]] = None


class PasswordResetByAdmin(BaseModel):
    new_password: str


def _user_to_out(u: Dict[str, Any]) -> Dict[str, Any]:
    return {k: v for k, v in u.items() if k not in ("_id", "password_hash")}


@api.get("/users")
async def list_users(_=Depends(require_cap("manage_users"))):
    docs = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return docs


@api.get("/users/roles")
async def role_info(_=Depends(require_cap("manage_users"))):
    return {
        "roles": [ROLE_ADMIN, ROLE_AUTHOR, ROLE_GUEST],
        "default_permissions": ROLE_DEFAULT_PERMS,
        "all_pages": ALL_PAGES,
        "max_authors": MAX_AUTHORS,
    }


@api.post("/users")
async def create_user(body: UserIn, _=Depends(require_cap("manage_users"))):
    if body.role not in (ROLE_ADMIN, ROLE_AUTHOR, ROLE_GUEST):
        raise HTTPException(400, "Invalid role")
    if body.role == ROLE_AUTHOR:
        count = await db.users.count_documents({"role": ROLE_AUTHOR})
        if count >= MAX_AUTHORS:
            raise HTTPException(400, f"Maximum {MAX_AUTHORS} Author users are allowed.")
    existing = await db.users.find_one({"email": body.email})
    if existing:
        raise HTTPException(400, "Email already registered")
    perms = body.page_permissions if body.page_permissions is not None else ROLE_DEFAULT_PERMS.get(body.role, [])
    doc = {
        "id": str(uuid.uuid4()),
        "email": body.email,
        "full_name": body.full_name,
        "mobile": body.mobile or "",
        "password_hash": hash_password(body.password),
        "role": body.role,
        "status": body.status if body.status in ("active", "inactive") else "active",
        "page_permissions": perms,
        "created_at": now_iso(),
        "last_login": None,
    }
    await db.users.insert_one(dict(doc))
    return _user_to_out(doc)


@api.put("/users/{user_id}")
async def update_user(user_id: str, body: UserUpdate, admin=Depends(require_cap("manage_users"))):
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(404, "User not found")
    updates: Dict[str, Any] = {}
    if body.full_name is not None:
        updates["full_name"] = body.full_name
    if body.mobile is not None:
        updates["mobile"] = body.mobile
    if body.role is not None:
        if body.role not in (ROLE_ADMIN, ROLE_AUTHOR, ROLE_GUEST):
            raise HTTPException(400, "Invalid role")
        if body.role == ROLE_AUTHOR and target.get("role") != ROLE_AUTHOR:
            count = await db.users.count_documents({"role": ROLE_AUTHOR})
            if count >= MAX_AUTHORS:
                raise HTTPException(400, f"Maximum {MAX_AUTHORS} Author users are allowed.")
        if target["email"] == ADMIN_EMAIL and body.role != ROLE_ADMIN:
            raise HTTPException(400, "Cannot change role of the primary admin")
        updates["role"] = body.role
        if body.page_permissions is None:
            updates["page_permissions"] = ROLE_DEFAULT_PERMS.get(body.role, [])
    if body.status is not None:
        if target["email"] == ADMIN_EMAIL and body.status != "active":
            raise HTTPException(400, "Cannot deactivate primary admin")
        updates["status"] = body.status
    if body.page_permissions is not None:
        updates["page_permissions"] = body.page_permissions
    if updates:
        await db.users.update_one({"id": user_id}, {"$set": updates})
    doc = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return doc


@api.delete("/users/{user_id}")
async def delete_user(user_id: str, admin=Depends(require_cap("manage_users"))):
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(404, "User not found")
    if target["email"] == ADMIN_EMAIL:
        raise HTTPException(400, "Cannot delete primary admin")
    await db.users.delete_one({"id": user_id})
    return {"ok": True}


@api.post("/users/{user_id}/reset-password")
async def admin_reset_password(user_id: str, body: PasswordResetByAdmin, _=Depends(require_cap("manage_users"))):
    if len(body.new_password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(404, "User not found")
    await db.users.update_one({"id": user_id}, {"$set": {"password_hash": hash_password(body.new_password)}})
    return {"ok": True}


# ---------- Bulk WhatsApp ----------
@api.get("/whatsapp/bulk-pending")
async def bulk_pending(fy: Optional[str] = None, admin=Depends(get_current_admin)):
    """Build a list of {phone, message} entries for all overdue students. Client opens wa.me sequentially with delay."""
    docs = await db.students.find({}, {"_id": 0}).to_list(10000)
    if fy:
        docs = [d for d in docs if _in_fy(d.get("admission_date"), fy)]
    out = []
    today = datetime.now(timezone.utc)
    for d in docs:
        full = await student_to_out(d)
        if full["status"] == "completed":
            continue
        nd = _parse_dt(full.get("next_due_date") or full.get("due_date"))
        if not nd or nd > today:
            continue
        out.append({
            "student_id": full["id"],
            "student_name": full["name"],
            "school_name": full["school_name"],
            "parent_name": full["parent_name"],
            "phone": full["parent_mobile"],
            "pending_amount": full["pending_amount"],
            "next_due_date": full.get("next_due_date") or full.get("due_date"),
            "overdue_days": full["overdue_days"],
        })
    return {"count": len(out), "items": out}


# ---------- Archive (Firebase Storage) ----------
@api.get("/archive/status")
async def archive_status(_=Depends(require_cap("archive"))):
    bucket = _get_bucket()
    archives = await db.archives.find({}, {"_id": 0}).to_list(100)
    return {
        "firebase_ready": bucket is not None,
        "bucket": FIREBASE_BUCKET if bucket is not None else None,
        "archives": archives,
    }


@api.post("/archive/backup")
async def archive_backup(fy: str = Query(..., description="Financial year label e.g. 2026-2027"), _=Depends(require_cap("archive"))):
    import json as _json
    bucket = _get_bucket()
    schools = await db.schools.find({}, {"_id": 0}).to_list(10000)
    students = await db.students.find({}, {"_id": 0}).to_list(50000)
    students = [s for s in students if _in_fy(s.get("admission_date"), fy)]
    sids = [s["id"] for s in students]
    payments = await db.payments.find({"student_id": {"$in": sids}}, {"_id": 0}).to_list(100000) if sids else []
    payments = [p for p in payments if _in_fy(p.get("payment_date"), fy)]
    snapshot = {
        "fy": fy,
        "exported_at": now_iso(),
        "schools": schools,
        "students": students,
        "payments": payments,
        "counts": {"schools": len(schools), "students": len(students), "payments": len(payments)},
    }
    payload = _json.dumps(snapshot, default=str).encode("utf-8")
    storage_url = None
    storage_error: Optional[str] = None
    if bucket is not None:
        try:
            blob = bucket.blob(f"archives/busfee-{fy}.json")
            blob.upload_from_string(payload, content_type="application/json")
            try:
                storage_url = blob.public_url
            except Exception:
                storage_url = f"gs://{FIREBASE_BUCKET}/archives/busfee-{fy}.json"
        except Exception as exc:
            storage_error = str(exc).splitlines()[0][:200]
            logger.warning(f"Firebase upload failed, falling back to MongoDB: {storage_error}")
    await db.archives.update_one(
        {"fy": fy},
        {"$set": {
            "fy": fy, "exported_at": snapshot["exported_at"],
            "counts": snapshot["counts"], "storage_url": storage_url,
            "size_bytes": len(payload),
            "local_snapshot": snapshot if storage_url is None else None,
        }},
        upsert=True,
    )
    return {
        "ok": True, "fy": fy, "counts": snapshot["counts"], "storage_url": storage_url,
        "stored_in": "firebase" if storage_url else "mongodb-local",
        "warning": storage_error,
    }


@api.post("/archive/restore")
async def archive_restore(fy: str = Query(...), _=Depends(require_cap("archive"))):
    import json as _json
    bucket = _get_bucket()
    arch = await db.archives.find_one({"fy": fy}, {"_id": 0})
    if not arch:
        raise HTTPException(404, "No archive recorded for that FY")
    snapshot = None
    if bucket is not None and arch.get("storage_url"):
        blob = bucket.blob(f"archives/busfee-{fy}.json")
        if blob.exists():
            data = blob.download_as_bytes()
            snapshot = _json.loads(data.decode("utf-8"))
    if snapshot is None:
        snapshot = arch.get("local_snapshot")
    if not snapshot:
        raise HTTPException(404, "Archive payload not available")

    # idempotent restore: upsert by id, do not duplicate
    for s in snapshot.get("schools", []):
        await db.schools.update_one({"id": s["id"]}, {"$set": s}, upsert=True)
    for st in snapshot.get("students", []):
        await db.students.update_one({"id": st["id"]}, {"$set": st}, upsert=True)
    for p in snapshot.get("payments", []):
        await db.payments.update_one({"id": p["id"]}, {"$set": p}, upsert=True)
    return {"ok": True, "fy": fy, "restored": snapshot.get("counts", {})}


# ---------- Health ----------
@api.get("/")
async def root():
    return {"service": "school-bus-fee-management", "status": "ok"}


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)
